import os
import threading
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from resume import extract_resume_data

from dotenv import load_dotenv
load_dotenv()

app = FastAPI(title="Resume Extractor")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Resolve paths relative to this file (so it works from any CWD)
BASE_DIR = Path(__file__).resolve().parent          # .../backend
PROJECT_ROOT = BASE_DIR.parent                       # .../your-project
FRONTEND_DIR = PROJECT_ROOT / "frontend"
EXCEL_PATH = os.getenv("EXCEL_PATH", str(PROJECT_ROOT / "resumes.xlsx"))

SUPPORTED_EXTS = {".pdf", ".docx", ".doc", ".png", ".jpg", ".jpeg",
                  ".tiff", ".tif", ".bmp", ".heif"}

excel_lock = threading.Lock()


@app.post("/api/upload")
async def upload_resume(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in SUPPORTED_EXTS:
        raise HTTPException(400, f"Unsupported file type: {ext}")

    content = await file.read()
    try:
        with excel_lock:
            row = extract_resume_data(content, file.filename, EXCEL_PATH)
        return {"success": True, "data": row}
    except Exception as e:
        raise HTTPException(500, str(e))

from openpyxl import load_workbook

@app.get("/api/data")
def get_data():
    if not os.path.exists(EXCEL_PATH):
        return {"headers": [], "rows": []}
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {"headers": [], "rows": []}
    headers = list(all_rows[0])
    # row_idx is the 1-indexed Excel row (header=1, first data row=2)
    data = [{"row_idx": i + 2, "cells": list(r)} for i, r in enumerate(all_rows[1:])]
    data.reverse()  # newest first
    return {"headers": headers, "rows": data}


@app.delete("/api/data/{row_idx}")
def delete_row(row_idx: int):
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(404, "Excel file not found")
    if row_idx < 2:
        raise HTTPException(400, "Cannot delete header row")
    with excel_lock:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        if row_idx > ws.max_row:
            raise HTTPException(404, "Row not found")
        ws.delete_rows(row_idx)
        wb.save(EXCEL_PATH)
    return {"success": True}

@app.get("/api/download")
def download_excel():
    if not os.path.exists(EXCEL_PATH):
        raise HTTPException(404, "No resumes uploaded yet")
    return FileResponse(EXCEL_PATH, filename="resumes.xlsx")


app.mount("/", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")